import openpyxl, os, json, shutil
from datetime import datetime

class Excel_Editor:
    def __init__(self, filename: str, sheet_name: str = None) -> None:
        """
        Initialize an Excel Editor.
        Auto create a new excel file if the file is not exists 
        Auto select the first sheet if the sheet name is not provided

        Parameters:
        filename (Any): The path to the Excel file to be edited.
        sheet_name (Any, optional): The name of the sheet to edit. Defaults to None.

        Returns:
        None: This method initializes the Excel editor.

        Examples:
        >>> editor = Excel_Editor('file.xlsx', 'Sheet1')
        """
        if not os.path.exists(filename): self.create_new_excel_file(filename, sheet_name)
        self.workbook = openpyxl.load_workbook(filename)
        self.filename = filename
        if sheet_name and not sheet_name in self.workbook.sheetnames: self.create_new_sheet(sheet_name)
        self.sheet_name = sheet_name
        self.sheet = self.workbook.active if sheet_name is None else self.workbook[self.sheet_name]

    def create_new_excel_file(self, filename: str, sheet_name: str = None):
        print("Auto Create excel file")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name if sheet_name is not None else "Sheet1"
        workbook.save(filename)

    def fill_first_row(self, row):
        for i, value in enumerate(row, start=1):
            self.sheet.cell(row=1, column=i).value = value
        
        self.workbook.save(self.filename)
    
    def add_new_row(self, row):
        next_row = self.sheet.max_row + 1
        for col, value in enumerate(row, start=1):
            self.sheet.cell(row=next_row, column=col).value = value
        self.workbook.save(self.filename)

    def fill_block(self, row_index, title, data, cover=True):
        column_index = None
        for cell in self.sheet[1]:
            if cell.value == title:
                column_index = cell.column
                break
        if column_index is None:
            raise ValueError(f"Title '{title}' not found in the first row.")
        cell = self.sheet.cell(row=row_index, column=column_index)
        if cover: cell.value = data
        else: cell.value += data
        self.workbook.save(self.filename)
    
    def find_row_index(self, values):
        for row in self.sheet.iter_rows(min_row=2):
            row_values = [cell.value for cell in row]
            if set(values).issubset(set(row_values)):
                return row[0].row
        return None

    def create_new_sheet(self, sheet_name):
        if sheet_name in self.workbook.sheetnames:
            print(f"{sheet_name} cant create because there already have a sheet call {sheet_name}")
        else:
            self.workbook.create_sheet(sheet_name)
            self.workbook.save(self.filename)
    
    def rename_sheet(self, new_name):
        self.sheet.title = new_name
        self.workbook.save(self.filename)

    def delete_sheet(self, sheet_name: str):
        """Delete a specified sheet from the workbook."""
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])
            self.workbook.save(self.filename)
        else:
            print(f"Sheet '{sheet_name}' does not exist.")

    def backup(self, backup_folder: str = "", max_backups: int = 5):
        """Create a backup of the current Excel file in a specified folder."""
        if not backup_folder: backup_folder = f"{os.path.basename(self.filename).replace('.xlsx', '')}_backups"
        os.makedirs(backup_folder, exist_ok=True)
        backup_filename = os.path.join(backup_folder, f"{os.path.basename(self.filename).replace('.xlsx', '')}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copyfile(self.filename, backup_filename)
        print(f"Backup created: {backup_filename}")
        self.manage_backups(backup_folder, max_backups)

    def manage_backups(self, backup_folder: str, max_backups: int):
        backups = sorted([f for f in os.listdir(backup_folder) if f.startswith(os.path.basename(self.filename.replace('.xlsx', '')) + '_backup_')])
        while len(backups) > max_backups:
            oldest_backup = os.path.join(backup_folder, backups[0])
            os.remove(oldest_backup)
            print(f"Deleted old backup: {oldest_backup}")
            backups.pop(0)

class Json_Editor:
    def __init__(self, filename):
        """
        Initialize an Json Editor.
        Auto create a new json file if the file is not exists 

        Parameters:
        filename (Any): The path to the Excel file to be edited.

        Returns:
        None: This method initializes the Excel editor.

        Examples:
        >>> editor = Json_Editor('file.json')
        """
        self.filename = filename
        if not os.path.exists(filename): self.create_new_json_file(filename)
        with open(self.filename, 'r') as json_file:
                self.data = json.load(json_file)

    def create_new_json_file(self, filename: str):
        print("Auto Create json file")
        with open(filename, "w") as json_file:
            json.dump("{}", json_file, indent=4)
        
    def add_new_key(self, key, value=""):
        """Add a new key-value pair to the JSON data."""
        self.data[key] = value
        with open(self.filename, 'w') as json_file:
            json.dump(self.data, json_file, indent=4)
    
    def update_key(self, key, value, cover=True):
        if key not in self.data:
            self.add_new_key(key, value)
        else:
            if cover: self.data[key] = value
            else: self.data[key] += value 
        with open(self.filename, 'w') as json_file:
            json.dump(self.data, json_file, indent=4)

    def all_keys(self):
        return self.data.keys()

    def read(self):
        return self.data
    
    def backup(self, backup_folder: str = "", max_backups: int = 5):
        """Create a backup of the current Json file in a specified folder."""
        if not backup_folder: backup_folder = f"{os.path.basename(self.filename).replace('.json', '')}_backups"
        os.makedirs(backup_folder, exist_ok=True)
        backup_filename = os.path.join(backup_folder, f"{os.path.basename(self.filename).replace('.json', '')}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        shutil.copyfile(self.filename, backup_filename)
        print(f"Backup created: {backup_filename}")
        self.manage_backups(backup_folder, max_backups)

    def manage_backups(self, backup_folder: str, max_backups: int):
        backups = sorted([f for f in os.listdir(backup_folder) if f.startswith(os.path.basename(self.filename.replace('.json', '')) + '_backup_')])
        while len(backups) > max_backups:
            oldest_backup = os.path.join(backup_folder, backups[0])
            os.remove(oldest_backup)
            print(f"Deleted old backup: {oldest_backup}")
            backups.pop(0)
